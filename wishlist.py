import json
import time
import requests
import datetime as dt
import openpyxl
import re

excel_file = "D:\Google Drive\Documents\My Notes\Entertainment\Game and Movie Releases.xlsx"
wb = openpyxl.load_workbook(excel_file)
sheet = wb['Steam Wishlist']


id = 'concretesurfer'
url = f'https://store.steampowered.com/wishlist/id/{id}/#sort=order'
wishlist_url =  json.loads( re.findall(r'g_strWishlistBaseURL = (".*?");', requests.get(url).text)[0] )


def Create_Column_Index(column_position = 1):
    column_index = {}
    for i in range(1, len(sheet[column_position])):
        main_column = sheet.cell(row=column_position, column=0+i).value
        if main_column is not None:
            column_index[main_column] = i
    return column_index


def Create_Row_Index(start=1, row_name = 'Name', row_position = 1):
    column_index = Create_Column_Index()
    row_index = {}
    for i in range(start, len(sheet['A'])):
        row = sheet.cell(row=i+1, column=column_index[row_name]).value
        if row is not None:
            row_index[row] = i+1
    return row_index


def UTC_Convert(utc):
    time_list = time.localtime(int(utc))
    converted_time = dt.datetime(time_list[0], time_list[1], time_list[2], time_list[3], time_list[4])
    return converted_time.strftime("%m-%d-%y")

    sheet.cell(row=row_index[row], column=column_index[column]).value = value


def Get_Wishlist():
    column_index = Create_Column_Index()
    row_index = Create_Row_Index()
    page = 0
    start_pos = 2
    data = 'start'
    while len(data) > 0:
        data = requests.get(f'{wishlist_url}wishlistdata/?p={page}').json()
        for key in data:
            name = data[key]['name']
            try:
                price = str(data[key]['subs'][0]['price'])
                formatted_price = float(f'{price[0:-2]}.{price[-2:]}')
            except IndexError:
                formatted_price = 'Not Listed'
            release_string = data[key]['release_string']
            release_date = UTC_Convert(data[key]['release_date'])
            print(f'{name} releases on {release_date}.')
            # if int(dt.datetime.now()) < int(converted_time):
            #     print('Game is not out yet.')
            sheet.cell(row=start_pos, column=column_index['Name']).value = name
            sheet.cell(row=start_pos, column=column_index['Price']).value = formatted_price
            sheet.cell(row=start_pos, column=column_index['Release Date']).value = release_date
            sheet.cell(row=start_pos, column=column_index['Release Info']).value = release_string
            start_pos +=1
        page += 1
        time.sleep(1)
        print(f'Total Pages: {page+1}')
    wb.save(excel_file)
    print('Complete')


if __name__ == "__main__":
    Get_Wishlist()
